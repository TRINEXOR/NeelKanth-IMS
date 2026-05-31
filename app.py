from flask import Flask, render_template, request, jsonify, send_file, Response, stream_with_context, session, redirect, url_for
import sqlite3, os, csv, io, json, time, sys, platform, hashlib, secrets
from datetime import datetime
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

DATA_DIR = os.environ.get('DATA_DIR', os.path.dirname(__file__))
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH = os.path.join(DATA_DIR, 'inventory.db')

# ─────────────────────────────────────────────
# Database
# ─────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute('''CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT 'General', quantity INTEGER NOT NULL DEFAULT 0,
            unit TEXT NOT NULL DEFAULT 'pcs', low_stock INTEGER NOT NULL DEFAULT 10,
            price REAL NOT NULL DEFAULT 0.0, supplier TEXT, notes TEXT,
            barcode TEXT, updated_at TEXT NOT NULL)''')

        conn.execute('''CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT, item_id INTEGER,
            item_name TEXT, action TEXT, details TEXT, timestamp TEXT,
            user TEXT DEFAULT 'system')''')

        conn.execute('''CREATE TABLE IF NOT EXISTS alert_config (
            id INTEGER PRIMARY KEY CHECK (id=1),
            enabled INTEGER NOT NULL DEFAULT 1,
            last_notified TEXT)''')

        # ── NEW: users table ──────────────────────────────────────────────
        conn.execute('''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'viewer',
            created_at TEXT NOT NULL)''')

        # ── NEW: theme preferences ────────────────────────────────────────
        conn.execute('''CREATE TABLE IF NOT EXISTS user_prefs (
            user_id INTEGER PRIMARY KEY,
            theme TEXT NOT NULL DEFAULT 'dark',
            FOREIGN KEY(user_id) REFERENCES users(id))''')

        conn.execute("INSERT OR IGNORE INTO alert_config (id,enabled) VALUES (1,1)")
        _migrate_tables(conn)
        _seed_admin(conn)

        cur = conn.execute('SELECT COUNT(*) FROM items')
        if cur.fetchone()[0] == 0:
            seed = [
                ('Rice (Basmati)','Grains',    120,'kg',  20,2.50,'FoodCorp',  'Long grain', None),
                ('Wheat Flour',  'Grains',       8,'kg',  15,1.80,'MillersInc','All-purpose', None),
                ('Cooking Oil',  'Oils',          5,'L',   10,3.20,'PureOils',  None, None),
                ('Canned Tomato','Canned Goods', 42,'cans',15,0.90,'CanCo',     None, None),
                ('Lentils',      'Legumes',       3,'kg',  10,2.10,'FoodCorp',  'Red lentils', None),
                ('Sugar',        'Condiments',   30,'kg',  10,1.20,'SugarMill', None, None),
                ('Salt',         'Condiments',   25,'kg',   5,0.40,'SaltWorks', 'Iodised', None),
                ('Pasta',        'Grains',        7,'packs',10,1.60,'PastaCo',  '500g packs', None),
                ('Baby Formula', 'Baby',          2,'cans', 5,12.00,'NutriKids','Stage 2', None),
                ('Soap Bars',    'Hygiene',      18,'bars', 8,0.70,'CleanCo',   None, None),
            ]
            now = datetime.now().isoformat(timespec='seconds')
            for r in seed:
                conn.execute('INSERT INTO items (name,category,quantity,unit,low_stock,price,supplier,notes,barcode,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?)', (*r, now))
        conn.commit()

def _seed_admin(conn):
    """Create default admin user if no users exist."""
    count = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
    if count == 0:
        now = datetime.now().isoformat(timespec='seconds')
        pwd = _hash_password('admin123')
        conn.execute('INSERT INTO users (username,password_hash,role,created_at) VALUES (?,?,?,?)',
                     ('admin', pwd, 'admin', now))

def _migrate_tables(conn):
    """Add missing columns to existing tables for backward compatibility."""
    # alert_config migration
    cols = {r[1] for r in conn.execute('PRAGMA table_info(alert_config)').fetchall()}
    if 'last_notified' not in cols:
        conn.execute('ALTER TABLE alert_config ADD COLUMN last_notified TEXT')

    # items: add barcode column if missing
    item_cols = {r[1] for r in conn.execute('PRAGMA table_info(items)').fetchall()}
    if 'barcode' not in item_cols:
        conn.execute('ALTER TABLE items ADD COLUMN barcode TEXT')

    # activity_log: add user column if missing
    log_cols = {r[1] for r in conn.execute('PRAGMA table_info(activity_log)').fetchall()}
    if 'user' not in log_cols:
        conn.execute('ALTER TABLE activity_log ADD COLUMN user TEXT DEFAULT "system"')

# ─────────────────────────────────────────────
# Auth Helpers
# ─────────────────────────────────────────────
def _hash_password(password):
    salt = 'nkims_salt_v1'
    return hashlib.sha256(f'{salt}{password}'.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Authentication required', 'login': True}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated

def current_user():
    return session.get('username', 'system')

# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
def get_low_stock_items():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM items WHERE quantity <= low_stock ORDER BY quantity ASC').fetchall()
    return [dict(r) for r in rows]

def log_action(conn, item_id, item_name, action, details):
    now = datetime.now().isoformat(timespec='seconds')
    conn.execute('INSERT INTO activity_log VALUES (NULL,?,?,?,?,?,?)',
                 (item_id, item_name, action, details, now, current_user()))

# ─────────────────────────────────────────────
# Auth Routes
# ─────────────────────────────────────────────
@app.route('/login')
def login_page():
    if 'user_id' in session:
        return redirect(url_for('index'))
    return render_template('index.html', show_login=True)

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json()
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    with get_db() as conn:
        user = conn.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
    if not user or user['password_hash'] != _hash_password(password):
        return jsonify({'error': 'Invalid username or password'}), 401
    session.permanent = True
    session['user_id'] = user['id']
    session['username'] = user['username']
    session['role'] = user['role']
    # Load theme pref
    with get_db() as conn:
        pref = conn.execute('SELECT theme FROM user_prefs WHERE user_id=?', (user['id'],)).fetchone()
    theme = pref['theme'] if pref else 'dark'
    return jsonify({'message': 'Login successful', 'username': user['username'], 'role': user['role'], 'theme': theme})

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'message': 'Logged out'})

@app.route('/api/auth/me')
def api_me():
    if 'user_id' not in session:
        return jsonify({'logged_in': False})
    with get_db() as conn:
        pref = conn.execute('SELECT theme FROM user_prefs WHERE user_id=?', (session['user_id'],)).fetchone()
    theme = pref['theme'] if pref else 'dark'
    return jsonify({'logged_in': True, 'username': session.get('username'), 'role': session.get('role'), 'theme': theme})

@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.get_json()
    old_pw = data.get('old_password', '')
    new_pw = data.get('new_password', '')
    if len(new_pw) < 6:
        return jsonify({'error': 'New password must be at least 6 characters'}), 400
    with get_db() as conn:
        user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
        if user['password_hash'] != _hash_password(old_pw):
            return jsonify({'error': 'Current password is incorrect'}), 401
        conn.execute('UPDATE users SET password_hash=? WHERE id=?', (_hash_password(new_pw), session['user_id']))
        conn.commit()
    return jsonify({'message': 'Password changed successfully'})

# ── Admin: user management ────────────────────
@app.route('/api/admin/users', methods=['GET'])
@login_required
@admin_required
def list_users():
    with get_db() as conn:
        rows = conn.execute('SELECT id,username,role,created_at FROM users ORDER BY id').fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/admin/users', methods=['POST'])
@login_required
@admin_required
def create_user():
    data = request.get_json()
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    role = data.get('role', 'viewer')
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    if role not in ('admin', 'editor', 'viewer'):
        return jsonify({'error': 'Role must be admin, editor, or viewer'}), 400
    now = datetime.now().isoformat(timespec='seconds')
    try:
        with get_db() as conn:
            conn.execute('INSERT INTO users (username,password_hash,role,created_at) VALUES (?,?,?,?)',
                         (username, _hash_password(password), role, now))
            conn.commit()
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Username already exists'}), 409
    return jsonify({'message': f'User {username} created'}), 201

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_user(user_id):
    if user_id == session['user_id']:
        return jsonify({'error': 'Cannot delete yourself'}), 400
    with get_db() as conn:
        conn.execute('DELETE FROM users WHERE id=?', (user_id,))
        conn.commit()
    return jsonify({'message': 'User deleted'})

# ─────────────────────────────────────────────
# Theme Preference
# ─────────────────────────────────────────────
@app.route('/api/prefs/theme', methods=['POST'])
@login_required
def save_theme():
    data = request.get_json()
    theme = data.get('theme', 'dark')
    if theme not in ('dark', 'light'):
        return jsonify({'error': 'Invalid theme'}), 400
    with get_db() as conn:
        conn.execute('INSERT OR REPLACE INTO user_prefs (user_id, theme) VALUES (?,?)',
                     (session['user_id'], theme))
        conn.commit()
    return jsonify({'theme': theme})

# ─────────────────────────────────────────────
# Core Routes
# ─────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/print')
@login_required
def print_view():
    """Clean printable inventory page."""
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM items ORDER BY category, name').fetchall()
        stats = conn.execute(
            'SELECT COUNT(*) as t, COALESCE(SUM(quantity*price),0) as v, '
            'SUM(CASE WHEN quantity<=low_stock THEN 1 ELSE 0 END) as l FROM items'
        ).fetchone()
    items = []
    for r in rows:
        d = dict(r)
        d['is_low'] = d['quantity'] <= d['low_stock']
        d['is_out'] = d['quantity'] == 0
        d['stock_value'] = round(d['quantity'] * d['price'], 2)
        items.append(d)
    return render_template('print.html', items=items, stats=dict(stats),
                           generated=datetime.now().strftime('%d %b %Y %H:%M'))

# ─────────────────────────────────────────────
# Core CRUD
# ─────────────────────────────────────────────
@app.route('/api/items')
@login_required
def get_items():
    search   = request.args.get('search','').strip()
    category = request.args.get('category','').strip()
    low_only = request.args.get('low_only','false') == 'true'
    sort_by  = request.args.get('sort','name')
    order    = request.args.get('order','asc')
    allowed  = {'name','category','quantity','price','updated_at'}
    if sort_by not in allowed: sort_by = 'name'
    order = 'DESC' if order == 'desc' else 'ASC'
    q = 'SELECT * FROM items WHERE 1=1'; p = []
    if search:
        q += ' AND (name LIKE ? OR supplier LIKE ? OR notes LIKE ? OR barcode LIKE ?)'
        p += [f'%{search}%']*4
    if category: q += ' AND category = ?'; p.append(category)
    if low_only: q += ' AND quantity <= low_stock'
    q += f' ORDER BY {sort_by} {order}'
    with get_db() as conn:
        rows = conn.execute(q, p).fetchall()
    items = []
    for r in rows:
        d = dict(r)
        d['is_low'] = d['quantity'] <= d['low_stock']
        d['is_out'] = d['quantity'] == 0
        items.append(d)
    return jsonify(items)

@app.route('/api/stats')
@login_required
def get_stats():
    with get_db() as conn:
        total     = conn.execute('SELECT COUNT(*) FROM items').fetchone()[0]
        low_stock = conn.execute('SELECT COUNT(*) FROM items WHERE quantity<=low_stock AND quantity>0').fetchone()[0]
        out_stock = conn.execute('SELECT COUNT(*) FROM items WHERE quantity=0').fetchone()[0]
        value     = conn.execute('SELECT COALESCE(SUM(quantity*price),0) FROM items').fetchone()[0]
        cats      = conn.execute('SELECT COUNT(DISTINCT category) FROM items').fetchone()[0]
    return jsonify({'total':total,'low_stock':low_stock,'out_stock':out_stock,'total_value':round(value,2),'categories':cats})

@app.route('/api/categories')
@login_required
def get_categories():
    with get_db() as conn:
        rows = conn.execute('SELECT DISTINCT category FROM items ORDER BY category').fetchall()
    return jsonify([r[0] for r in rows])

@app.route('/api/log')
@login_required
def get_log():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM activity_log ORDER BY id DESC LIMIT 100').fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/items', methods=['POST'])
@login_required
def create_item():
    if session.get('role') == 'viewer':
        return jsonify({'error': 'Viewers cannot add items'}), 403
    data = request.get_json()
    for f in ['name','category','quantity','unit','low_stock','price']:
        if f not in data: return jsonify({'error': f'Missing: {f}'}), 400
    now = datetime.now().isoformat(timespec='seconds')
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO items (name,category,quantity,unit,low_stock,price,supplier,notes,barcode,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?)',
            (data['name'],data['category'],int(data['quantity']),data['unit'],
             int(data['low_stock']),float(data['price']),data.get('supplier',''),
             data.get('notes',''),data.get('barcode',''),now))
        log_action(conn, cur.lastrowid, data['name'], 'Added', f"Qty: {data['quantity']} {data['unit']}")
        conn.commit()
    return jsonify({'id':cur.lastrowid,'message':'Item created'}), 201

@app.route('/api/items/<int:item_id>', methods=['PUT'])
@login_required
def update_item(item_id):
    if session.get('role') == 'viewer':
        return jsonify({'error': 'Viewers cannot edit items'}), 403
    data = request.get_json()
    now = datetime.now().isoformat(timespec='seconds')
    with get_db() as conn:
        e = conn.execute('SELECT * FROM items WHERE id=?',(item_id,)).fetchone()
        if not e: return jsonify({'error':'Not found'}), 404
        conn.execute(
            'UPDATE items SET name=?,category=?,quantity=?,unit=?,low_stock=?,price=?,supplier=?,notes=?,barcode=?,updated_at=? WHERE id=?',
            (data.get('name',e['name']),data.get('category',e['category']),
             int(data.get('quantity',e['quantity'])),data.get('unit',e['unit']),
             int(data.get('low_stock',e['low_stock'])),float(data.get('price',e['price'])),
             data.get('supplier',e['supplier']),data.get('notes',e['notes']),
             data.get('barcode',e['barcode'] or ''),now,item_id))
        log_action(conn, item_id, data.get('name',e['name']), 'Updated',
                   f"Qty: {data.get('quantity',e['quantity'])}")
        conn.commit()
    return jsonify({'message':'Item updated'})

@app.route('/api/items/<int:item_id>', methods=['DELETE'])
@login_required
def delete_item(item_id):
    if session.get('role') == 'viewer':
        return jsonify({'error': 'Viewers cannot delete items'}), 403
    with get_db() as conn:
        e = conn.execute('SELECT * FROM items WHERE id=?',(item_id,)).fetchone()
        if not e: return jsonify({'error':'Not found'}), 404
        conn.execute('DELETE FROM items WHERE id=?',(item_id,))
        log_action(conn, item_id, e['name'], 'Deleted', '—')
        conn.commit()
    return jsonify({'message':'Item deleted'})

# ─────────────────────────────────────────────
# FEATURE 1: Inline Quantity Edit
# ─────────────────────────────────────────────
@app.route('/api/items/<int:item_id>/quantity', methods=['PATCH'])
@login_required
def patch_quantity(item_id):
    """Fast single-field quantity update — used by inline table edit."""
    if session.get('role') == 'viewer':
        return jsonify({'error': 'Viewers cannot edit items'}), 403
    data = request.get_json()
    if 'quantity' not in data:
        return jsonify({'error': 'Missing quantity'}), 400
    qty = int(data['quantity'])
    if qty < 0:
        return jsonify({'error': 'Quantity cannot be negative'}), 400
    now = datetime.now().isoformat(timespec='seconds')
    with get_db() as conn:
        e = conn.execute('SELECT * FROM items WHERE id=?', (item_id,)).fetchone()
        if not e: return jsonify({'error': 'Not found'}), 404
        old_qty = e['quantity']
        conn.execute('UPDATE items SET quantity=?, updated_at=? WHERE id=?', (qty, now, item_id))
        diff = qty - old_qty
        sign = '+' if diff >= 0 else ''
        log_action(conn, item_id, e['name'], 'Qty Adjusted',
                   f"{old_qty} → {qty} ({sign}{diff})")
        conn.commit()
    return jsonify({'message': 'Quantity updated', 'quantity': qty,
                    'is_low': qty <= e['low_stock'], 'is_out': qty == 0})

# ─────────────────────────────────────────────
# FEATURE 2: Bulk Stock Adjustment
# ─────────────────────────────────────────────
@app.route('/api/items/bulk-adjust', methods=['POST'])
@login_required
def bulk_adjust():
    """
    Adjust quantity for multiple items at once.
    Body: { "adjustments": [{"id": 1, "delta": 10}, {"id": 2, "delta": -5}] }
    Or set absolute:  { "adjustments": [{"id": 1, "quantity": 50}] }
    """
    if session.get('role') == 'viewer':
        return jsonify({'error': 'Viewers cannot edit items'}), 403
    data = request.get_json()
    adjustments = data.get('adjustments', [])
    if not adjustments:
        return jsonify({'error': 'No adjustments provided'}), 400
    now = datetime.now().isoformat(timespec='seconds')
    results = []
    with get_db() as conn:
        for adj in adjustments:
            item_id = adj.get('id')
            if not item_id: continue
            e = conn.execute('SELECT * FROM items WHERE id=?', (item_id,)).fetchone()
            if not e:
                results.append({'id': item_id, 'error': 'Not found'})
                continue
            old_qty = e['quantity']
            if 'quantity' in adj:
                new_qty = max(0, int(adj['quantity']))
                action_detail = f"{old_qty} → {new_qty} (set)"
            else:
                delta = int(adj.get('delta', 0))
                new_qty = max(0, old_qty + delta)
                sign = '+' if delta >= 0 else ''
                action_detail = f"{old_qty} → {new_qty} ({sign}{delta})"
            conn.execute('UPDATE items SET quantity=?, updated_at=? WHERE id=?',
                         (new_qty, now, item_id))
            log_action(conn, item_id, e['name'], 'Bulk Adjusted', action_detail)
            results.append({'id': item_id, 'name': e['name'],
                            'old_quantity': old_qty, 'new_quantity': new_qty,
                            'is_low': new_qty <= e['low_stock'], 'is_out': new_qty == 0})
        conn.commit()
    return jsonify({'updated': len(results), 'results': results})

# ─────────────────────────────────────────────
# FEATURE 3: Per-Item Stock History
# ─────────────────────────────────────────────
@app.route('/api/items/<int:item_id>/history')
@login_required
def item_history(item_id):
    """Full activity history for a single item."""
    with get_db() as conn:
        item = conn.execute('SELECT * FROM items WHERE id=?', (item_id,)).fetchone()
        if not item: return jsonify({'error': 'Not found'}), 404
        logs = conn.execute(
            'SELECT * FROM activity_log WHERE item_id=? ORDER BY id DESC LIMIT 200',
            (item_id,)
        ).fetchall()
    return jsonify({
        'item': dict(item),
        'history': [dict(r) for r in logs]
    })

# ─────────────────────────────────────────────
# FEATURE 4: Barcode Lookup
# ─────────────────────────────────────────────
@app.route('/api/items/barcode/<string:code>')
@login_required
def lookup_barcode(code):
    """Find an item by its barcode."""
    with get_db() as conn:
        row = conn.execute('SELECT * FROM items WHERE barcode=?', (code,)).fetchone()
    if not row:
        return jsonify({'error': 'No item found with that barcode'}), 404
    d = dict(row)
    d['is_low'] = d['quantity'] <= d['low_stock']
    d['is_out'] = d['quantity'] == 0
    return jsonify(d)

@app.route('/api/items/<int:item_id>/barcode', methods=['PATCH'])
@login_required
def set_barcode(item_id):
    """Assign or update a barcode for an item."""
    if session.get('role') == 'viewer':
        return jsonify({'error': 'Viewers cannot edit items'}), 403
    data = request.get_json()
    barcode = (data.get('barcode') or '').strip()
    now = datetime.now().isoformat(timespec='seconds')
    with get_db() as conn:
        e = conn.execute('SELECT * FROM items WHERE id=?', (item_id,)).fetchone()
        if not e: return jsonify({'error': 'Not found'}), 404
        if barcode:
            existing = conn.execute(
                'SELECT id FROM items WHERE barcode=? AND id!=?', (barcode, item_id)
            ).fetchone()
            if existing:
                return jsonify({'error': 'Barcode already assigned to another item'}), 409
        conn.execute('UPDATE items SET barcode=?, updated_at=? WHERE id=?',
                     (barcode or None, now, item_id))
        log_action(conn, item_id, e['name'], 'Barcode Set', barcode or 'cleared')
        conn.commit()
    return jsonify({'message': 'Barcode updated', 'barcode': barcode})

# ─────────────────────────────────────────────
# FEATURE 5: Restock Shopping List
# ─────────────────────────────────────────────
@app.route('/api/restock-list')
@login_required
def restock_list():
    """
    Returns all low/out-of-stock items grouped by supplier,
    with suggested reorder quantities (restock to 2× threshold).
    """
    with get_db() as conn:
        rows = conn.execute(
            '''SELECT * FROM items WHERE quantity <= low_stock ORDER BY supplier, name'''
        ).fetchall()
    items = []
    by_supplier = {}
    for r in rows:
        d = dict(r)
        d['is_out'] = d['quantity'] == 0
        d['suggested_order'] = max(0, d['low_stock'] * 2 - d['quantity'])
        d['estimated_cost'] = round(d['suggested_order'] * d['price'], 2)
        supplier = d['supplier'] or 'Unknown Supplier'
        by_supplier.setdefault(supplier, []).append(d)
        items.append(d)

    summary = []
    for supplier, sitems in sorted(by_supplier.items()):
        total_cost = sum(i['estimated_cost'] for i in sitems)
        summary.append({
            'supplier': supplier,
            'items': sitems,
            'total_items': len(sitems),
            'total_estimated_cost': round(total_cost, 2)
        })
    grand_total = sum(i['estimated_cost'] for i in items)
    return jsonify({
        'total_items': len(items),
        'grand_total_cost': round(grand_total, 2),
        'by_supplier': summary,
        'generated_at': datetime.now().isoformat(timespec='seconds')
    })

@app.route('/api/restock-list/export/csv')
@login_required
def export_restock_csv():
    """Download the restock list as a CSV."""
    with get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM items WHERE quantity <= low_stock ORDER BY supplier, name'
        ).fetchall()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['Supplier','Item','Category','Current Qty','Unit',
                     'Threshold','Suggested Order','Unit Price (₹)','Est. Cost (₹)','Notes'])
    for r in rows:
        suggested = max(0, r['low_stock'] * 2 - r['quantity'])
        cost = round(suggested * r['price'], 2)
        writer.writerow([r['supplier'] or 'Unknown', r['name'], r['category'],
                         r['quantity'], r['unit'], r['low_stock'],
                         suggested, r['price'], cost, r['notes'] or ''])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),
                     mimetype='text/csv', as_attachment=True,
                     download_name=f"restock_{datetime.now().strftime('%Y%m%d_%H%M')}.csv")

# ─────────────────────────────────────────────
# FEATURE 6: CSV / Excel Import
# ─────────────────────────────────────────────
@app.route('/api/import/csv', methods=['POST'])
@login_required
def import_csv():
    """
    Import items from a CSV file.
    Expected columns (case-insensitive): name, category, quantity, unit,
                                          low_stock, price, supplier, notes, barcode
    If an item with the same name already exists, it will be updated.
    """
    if session.get('role') == 'viewer':
        return jsonify({'error': 'Viewers cannot import data'}), 403
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.csv'):
        return jsonify({'error': 'Only CSV files are supported'}), 400
    try:
        content = f.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        # Normalize headers (lowercase, strip spaces)
        headers = [h.lower().strip().replace(' ', '_') for h in (reader.fieldnames or [])]
        if 'name' not in headers:
            return jsonify({'error': 'CSV must have a "name" column'}), 400
        now = datetime.now().isoformat(timespec='seconds')
        added = updated = skipped = 0
        errors = []
        with get_db() as conn:
            for i, raw_row in enumerate(reader, start=2):
                row = {k.lower().strip().replace(' ', '_'): v.strip() for k, v in raw_row.items() if k}
                name = row.get('name', '').strip()
                if not name:
                    skipped += 1
                    continue
                try:
                    qty       = int(row.get('quantity', 0) or 0)
                    low_stock = int(row.get('low_stock', row.get('low_stock_threshold', 10)) or 10)
                    price     = float(row.get('price', 0.0) or 0.0)
                except (ValueError, TypeError) as ex:
                    errors.append(f"Row {i} ({name}): {ex}")
                    skipped += 1
                    continue
                category = row.get('category', 'General') or 'General'
                unit     = row.get('unit', 'pcs') or 'pcs'
                supplier = row.get('supplier', '') or ''
                notes    = row.get('notes', '') or ''
                barcode  = row.get('barcode', '') or None
                existing = conn.execute('SELECT id FROM items WHERE name=?', (name,)).fetchone()
                if existing:
                    conn.execute(
                        'UPDATE items SET category=?,quantity=?,unit=?,low_stock=?,price=?,supplier=?,notes=?,barcode=?,updated_at=? WHERE id=?',
                        (category, qty, unit, low_stock, price, supplier, notes, barcode, now, existing['id']))
                    log_action(conn, existing['id'], name, 'Imported (updated)', f"Qty: {qty}")
                    updated += 1
                else:
                    cur = conn.execute(
                        'INSERT INTO items (name,category,quantity,unit,low_stock,price,supplier,notes,barcode,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?)',
                        (name, category, qty, unit, low_stock, price, supplier, notes, barcode, now))
                    log_action(conn, cur.lastrowid, name, 'Imported (new)', f"Qty: {qty}")
                    added += 1
            conn.commit()
        return jsonify({
            'message': f'Import complete: {added} added, {updated} updated, {skipped} skipped',
            'added': added, 'updated': updated, 'skipped': skipped, 'errors': errors
        })
    except Exception as ex:
        return jsonify({'error': f'Import failed: {str(ex)}'}), 500

@app.route('/api/import/template')
@login_required
def import_template():
    """Download a blank CSV template for importing."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['name','category','quantity','unit','low_stock','price','supplier','notes','barcode'])
    writer.writerow(['Example Item','General','100','pcs','10','5.00','My Supplier','Optional notes','123456789'])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),
                     mimetype='text/csv', as_attachment=True,
                     download_name='import_template.csv')

# ─────────────────────────────────────────────
# Analytics API (unchanged + extended)
# ─────────────────────────────────────────────
@app.route('/api/analytics')
@login_required
def get_analytics():
    with get_db() as conn:
        by_cat = conn.execute('''
            SELECT category,
                   COUNT(*) as item_count,
                   SUM(quantity) as total_qty,
                   ROUND(SUM(quantity*price),2) as stock_value
            FROM items GROUP BY category ORDER BY stock_value DESC''').fetchall()
        top_value = conn.execute('''
            SELECT name, category, quantity, price, ROUND(quantity*price,2) as total_value
            FROM items ORDER BY total_value DESC LIMIT 5''').fetchall()
        health = conn.execute('''
            SELECT
                SUM(CASE WHEN quantity=0 THEN 1 ELSE 0 END) as out_of_stock,
                SUM(CASE WHEN quantity>0 AND quantity<=low_stock THEN 1 ELSE 0 END) as low,
                SUM(CASE WHEN quantity>low_stock AND quantity<=low_stock*2 THEN 1 ELSE 0 END) as moderate,
                SUM(CASE WHEN quantity>low_stock*2 THEN 1 ELSE 0 END) as healthy
            FROM items''').fetchone()
        restock = conn.execute('''
            SELECT name, category, quantity, low_stock, unit,
                   ROUND((CAST(quantity AS REAL)/low_stock)*100,1) as pct
            FROM items WHERE quantity <= low_stock ORDER BY pct ASC LIMIT 8''').fetchall()
        trend = conn.execute('''
            SELECT DATE(timestamp) as day, COUNT(*) as ops
            FROM activity_log
            WHERE timestamp >= DATE('now','-14 days')
            GROUP BY day ORDER BY day''').fetchall()
    return jsonify({
        'by_category': [dict(r) for r in by_cat],
        'top_value_items': [dict(r) for r in top_value],
        'health': dict(health),
        'restock_urgency': [dict(r) for r in restock],
        'activity_trend': [dict(r) for r in trend],
    })

# ─────────────────────────────────────────────
# Export API (unchanged)
# ─────────────────────────────────────────────
@app.route('/api/export/csv')
@login_required
def export_csv():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM items ORDER BY category, name').fetchall()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['ID','Name','Category','Quantity','Unit','Low Stock Threshold','Price','Stock Value','Supplier','Notes','Barcode','Last Updated','Status'])
    for r in rows:
        status = 'Out of Stock' if r['quantity']==0 else ('Low Stock' if r['quantity']<=r['low_stock'] else 'OK')
        writer.writerow([r['id'],r['name'],r['category'],r['quantity'],r['unit'],r['low_stock'],
                         r['price'],round(r['quantity']*r['price'],2),r['supplier'] or '',
                         r['notes'] or '',r['barcode'] or '',r['updated_at'],status])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),
                     mimetype='text/csv', as_attachment=True,
                     download_name=f"inventory_{datetime.now().strftime('%Y%m%d_%H%M')}.csv")

@app.route('/api/export/excel')
@login_required
def export_excel():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM items ORDER BY category, name').fetchall()
        stats = conn.execute('SELECT COUNT(*) as t, SUM(quantity*price) as v, SUM(CASE WHEN quantity<=low_stock THEN 1 ELSE 0 END) as l FROM items').fetchone()

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Inventory"
    HDR_FILL  = PatternFill("solid", fgColor="1E2435")
    LOW_FILL  = PatternFill("solid", fgColor="3D2A00")
    OUT_FILL  = PatternFill("solid", fgColor="3D0000")
    OK_FILL   = PatternFill("solid", fgColor="0D2B20")
    HDR_FONT  = Font(name='Calibri', bold=True, color="A0B4D0", size=10)
    TITLE_FONT= Font(name='Calibri', bold=True, color="FFFFFF", size=14)
    thin   = Side(style='thin', color="2A3045")
    border = Border(left=thin,right=thin,top=thin,bottom=thin)

    ws.merge_cells('A1:M1')
    ws['A1'] = f"NeelKanth IMS — Inventory Report — {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = PatternFill("solid", fgColor="0F1117")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:M2')
    ws['A2'] = f"Total Items: {stats['t']}   |   Stock Value: ₹{round(stats['v'] or 0,2):,}   |   Low/Out of Stock: {stats['l']}"
    ws['A2'].font = Font(name='Calibri', color="6B7590", italic=True, size=9)
    ws['A2'].fill = PatternFill("solid", fgColor="181C27")
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    headers = ['ID','Name','Category','Quantity','Unit','Low Stock ≤','Price (₹)','Stock Value','Supplier','Notes','Barcode','Last Updated','Status']
    col_widths = [6,28,14,11,8,12,11,13,16,22,16,20,12]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22

    for ri, r in enumerate(rows, 4):
        status = 'Out of Stock' if r['quantity']==0 else ('Low Stock' if r['quantity']<=r['low_stock'] else 'OK')
        row_fill = OUT_FILL if status=='Out of Stock' else (LOW_FILL if status=='Low Stock' else OK_FILL)
        vals = [r['id'],r['name'],r['category'],r['quantity'],r['unit'],r['low_stock'],
                r['price'],round(r['quantity']*r['price'],2),r['supplier'] or '',
                r['notes'] or '',r['barcode'] or '',r['updated_at'],status]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = row_fill; cell.border = border
            cell.font = Font(name='Calibri', color="E8ECF4", size=9)
            cell.alignment = Alignment(vertical='center',
                                       horizontal='center' if ci in [1,4,5,6,7,8,13] else 'left')
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f"A3:M{3+len(rows)}"

    ws2 = wb.create_sheet("By Category")
    ws2['A1'] = "Category Summary"
    ws2['A1'].font = Font(name='Calibri', bold=True, color="FFFFFF", size=13)
    ws2['A1'].fill = PatternFill("solid", fgColor="0F1117")
    ws2.merge_cells('A1:E1'); ws2.row_dimensions[1].height = 28
    for ci, h in enumerate(['Category','Items','Total Qty','Stock Value (₹)','Avg Price (₹)'],1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.border = border
        cell.alignment = Alignment(horizontal='center')
    ws2.column_dimensions['A'].width = 16; ws2.column_dimensions['B'].width = 8
    ws2.column_dimensions['C'].width = 12; ws2.column_dimensions['D'].width = 16; ws2.column_dimensions['E'].width = 14
    with get_db() as conn:
        cat_rows = conn.execute('SELECT category,COUNT(*) as cnt,SUM(quantity) as qty,ROUND(SUM(quantity*price),2) as val,ROUND(AVG(price),2) as avg_p FROM items GROUP BY category ORDER BY val DESC').fetchall()
    for ri, r in enumerate(cat_rows, 3):
        for ci, v in enumerate([r['category'],r['cnt'],r['qty'],r['val'],r['avg_p']],1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.font = Font(name='Calibri', color="E8ECF4", size=9)
            cell.fill = PatternFill("solid", fgColor="181C27"); cell.border = border
            cell.alignment = Alignment(horizontal='center' if ci>1 else 'left', vertical='center')
        ws2.row_dimensions[ri].height = 18

    ws3 = wb.create_sheet("Low Stock")
    ws3['A1'] = "Low Stock Alert Report"
    ws3['A1'].font = Font(name='Calibri', bold=True, color="F95F5F", size=13)
    ws3['A1'].fill = PatternFill("solid", fgColor="0F1117")
    ws3.merge_cells('A1:G1'); ws3.row_dimensions[1].height = 28
    for ci, h in enumerate(['Name','Category','Current Qty','Unit','Threshold','% of Threshold','Urgency'],1):
        cell = ws3.cell(row=2, column=ci, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.border = border
        cell.alignment = Alignment(horizontal='center')
    ws3.column_dimensions['A'].width = 28; ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 13; ws3.column_dimensions['D'].width = 8
    ws3.column_dimensions['E'].width = 12; ws3.column_dimensions['F'].width = 16; ws3.column_dimensions['G'].width = 12
    low_items = [r for r in rows if r['quantity'] <= r['low_stock']]
    low_items.sort(key=lambda x: x['quantity']/max(x['low_stock'],1))
    for ri, r in enumerate(low_items, 3):
        pct = round((r['quantity']/max(r['low_stock'],1))*100,1)
        urgency = 'CRITICAL' if r['quantity']==0 else ('High' if pct<50 else 'Medium')
        row_fill = OUT_FILL if r['quantity']==0 else LOW_FILL
        for ci, v in enumerate([r['name'],r['category'],r['quantity'],r['unit'],r['low_stock'],f"{pct}%",urgency],1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.font = Font(name='Calibri', color="E8ECF4", size=9, bold=(ci==7))
            cell.fill = row_fill; cell.border = border
            cell.alignment = Alignment(horizontal='center' if ci>1 else 'left', vertical='center')
        ws3.row_dimensions[ri].height = 18

    for sheet in [ws, ws2, ws3]:
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.tabColor = "1E2435"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"inventory_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

# ─────────────────────────────────────────────
# Alert / Notification API (unchanged)
# ─────────────────────────────────────────────
@app.route('/api/alerts/config', methods=['GET'])
@login_required
def get_alert_config():
    with get_db() as conn:
        row = conn.execute('SELECT enabled, last_notified FROM alert_config WHERE id=1').fetchone()
    return jsonify(dict(row) if row else {'enabled': 1, 'last_notified': None})

@app.route('/api/alerts/config', methods=['POST'])
@login_required
def save_alert_config():
    data = request.get_json()
    with get_db() as conn:
        conn.execute('UPDATE alert_config SET enabled=? WHERE id=1',
                     (1 if data.get('enabled') else 0,))
        conn.commit()
    return jsonify({'message': 'Notification settings saved'})

@app.route('/api/alerts/low-stock')
@login_required
def get_alert_low_stock():
    items = get_low_stock_items()
    for i in items:
        i['is_low'] = i['quantity'] <= i['low_stock']
        i['is_out'] = i['quantity'] == 0
    return jsonify({'count': len(items), 'items': items})

@app.route('/api/alerts/ack', methods=['POST'])
@login_required
def ack_alert():
    now = datetime.now().isoformat(timespec='seconds')
    with get_db() as conn:
        conn.execute('UPDATE alert_config SET last_notified=? WHERE id=1', (now,))
        conn.commit()
    return jsonify({'last_notified': now})

# ─────────────────────────────────────────────
# FEATURE 7: Keyboard Shortcuts Reference
# ─────────────────────────────────────────────
@app.route('/api/shortcuts')
def get_shortcuts():
    """Returns all keyboard shortcuts for the frontend to display."""
    return jsonify([
        {'keys': ['Ctrl', 'N'],    'description': 'Add new item'},
        {'keys': ['/'],            'description': 'Focus search box'},
        {'keys': ['Ctrl', 'E'],    'description': 'Export CSV'},
        {'keys': ['Ctrl', 'P'],    'description': 'Open print view'},
        {'keys': ['Ctrl', 'B'],    'description': 'Open bulk adjust'},
        {'keys': ['Ctrl', 'R'],    'description': 'Open restock list'},
        {'keys': ['Ctrl', 'L'],    'description': 'Toggle light / dark mode'},
        {'keys': ['?'],            'description': 'Show this shortcuts guide'},
        {'keys': ['Esc'],          'description': 'Close any open modal'},
        {'keys': ['G', 'D'],       'description': 'Go to Dashboard'},
        {'keys': ['G', 'A'],       'description': 'Go to Analytics'},
        {'keys': ['G', 'L'],       'description': 'Go to Activity Log'},
        {'keys': ['G', 'N'],       'description': 'Go to Notifications'},
    ])

# ─────────────────────────────────────────────
# SSE Boot Sequence (updated with new features)
# ─────────────────────────────────────────────
def _sse(event, data):
    return f"event: {event}\ndata: {json.dumps(data)}\n\n"

@app.route('/api/boot-stream')
def boot_stream():
    def generate():
        yield "retry: 0\n\n"
        yield _sse('step', {'msg': 'Checking Python runtime', 'status': 'running'})
        time.sleep(0.35)
        py_ver = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
        yield _sse('step', {'msg': f'Python {py_ver} · {platform.system()} {platform.machine()}', 'status': 'ok'})
        time.sleep(0.2)
        yield _sse('step', {'msg': 'Starting Flask application', 'status': 'running'})
        time.sleep(0.3)
        yield _sse('step', {'msg': f'Flask app mounted · debug={app.debug}', 'status': 'ok'})
        time.sleep(0.2)
        yield _sse('step', {'msg': 'Locating SQLite database', 'status': 'running'})
        time.sleep(0.4)
        db_exists = os.path.exists(DB_PATH)
        db_size   = round(os.path.getsize(DB_PATH) / 1024, 1) if db_exists else 0
        yield _sse('step', {'msg': f'SQLite found · {os.path.basename(DB_PATH)} ({db_size} KB)', 'status': 'ok'})
        time.sleep(0.2)
        yield _sse('step', {'msg': 'Verifying database schema', 'status': 'running'})
        time.sleep(0.35)
        try:
            with get_db() as conn:
                tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
            yield _sse('step', {'msg': f'Schema OK · tables: {", ".join(tables)}', 'status': 'ok'})
        except Exception as e:
            yield _sse('step', {'msg': f'Schema error: {e}', 'status': 'error'})
        time.sleep(0.2)
        yield _sse('step', {'msg': 'Loading inventory records', 'status': 'running'})
        time.sleep(0.3)
        try:
            with get_db() as conn:
                item_count  = conn.execute('SELECT COUNT(*) FROM items').fetchone()[0]
                low_count   = conn.execute('SELECT COUNT(*) FROM items WHERE quantity<=low_stock').fetchone()[0]
                total_value = conn.execute('SELECT COALESCE(SUM(quantity*price),0) FROM items').fetchone()[0]
            yield _sse('step', {'msg': f'{item_count} items loaded · {low_count} low-stock · value ₹{round(total_value,2):,}', 'status': 'ok'})
        except Exception as e:
            yield _sse('step', {'msg': f'Inventory error: {e}', 'status': 'error'})
        time.sleep(0.2)
        yield _sse('step', {'msg': 'Checking user accounts', 'status': 'running'})
        time.sleep(0.3)
        try:
            with get_db() as conn:
                user_count = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
            yield _sse('step', {'msg': f'Auth ready · {user_count} user(s) registered', 'status': 'ok'})
        except Exception as e:
            yield _sse('step', {'msg': f'Auth error: {e}', 'status': 'error'})
        time.sleep(0.2)
        yield _sse('step', {'msg': 'Loading notification settings', 'status': 'running'})
        time.sleep(0.3)
        try:
            with get_db() as conn:
                cfg = conn.execute('SELECT enabled FROM alert_config WHERE id=1').fetchone()
                low_count = conn.execute('SELECT COUNT(*) FROM items WHERE quantity<=low_stock').fetchone()[0]
            if cfg and cfg['enabled']:
                yield _sse('step', {'msg': f'Browser notifications ON · {low_count} item(s) need attention', 'status': 'ok'})
            else:
                yield _sse('step', {'msg': 'Browser notifications disabled', 'status': 'ok'})
        except Exception as e:
            yield _sse('step', {'msg': f'Alert config error: {e}', 'status': 'error'})
        time.sleep(0.2)
        yield _sse('step', {'msg': 'Warming analytics engine', 'status': 'running'})
        time.sleep(0.4)
        try:
            with get_db() as conn:
                cat_count = conn.execute('SELECT COUNT(DISTINCT category) FROM items').fetchone()[0]
                log_count = conn.execute('SELECT COUNT(*) FROM activity_log').fetchone()[0]
            yield _sse('step', {'msg': f'Analytics ready · {cat_count} categories · {log_count} log entries', 'status': 'ok'})
        except Exception as e:
            yield _sse('step', {'msg': f'Analytics error: {e}', 'status': 'error'})
        time.sleep(0.2)
        yield _sse('step', {'msg': 'NeelKanth IMS · All systems operational', 'status': 'ok'})
        time.sleep(0.3)
        yield _sse('done', {'msg': 'Boot complete', 'timestamp': datetime.now().isoformat(timespec='seconds')})

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no', 'Access-Control-Allow-Origin': '*'}
    )

# ─────────────────────────────────────────────
init_db()

if __name__ == '__main__':
    if sys.platform == 'win32':
        sys.stdout.reconfigure(encoding='utf-8')
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', '1') == '1'
    print("\n🏪  NeelKanth IMS — Inventory Management System (Enhanced)")
    print("─" * 58)
    print(f"   Open:  http://127.0.0.1:{port}")
    print(f"   Login: admin / admin123  (change after first login!)")
    print("   Features: CRUD · Bulk Adjust · Inline Edit · Import/Export")
    print("             Analytics · Restock List · Barcode · Auth · Themes")
    print("─" * 58)
    app.run(host='0.0.0.0', port=port, debug=debug, threaded=True)
    
